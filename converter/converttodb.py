import sys
import os
sys.path.append(os.getcwd() + "/../classes")
import openpyxl
from airport import Airport
from routeappinfo import RouteAppInfo
import pickle
from connection import Connection
import math


def bubbleSort(connections): ##dont care how slow it is
    for passnum in range(len(connections)-1,0,-1):
        for i in range(passnum):
            if connections[i].getCode() >connections[i+1].getCode():
                temp = connections[i]
                connections[i] = connections[i+1]
                connections[i+1] = temp



airports = []


marketsWb = openpyxl.load_workbook('Markets.xlsx')
airInfoWb = openpyxl.load_workbook('airportdata.xlsx')

sheet = marketsWb.get_sheet_by_name('Sheet1')
print("Number of entries: " + str(sheet.max_row))

for i in range(2,sheet.max_row+1):
    apCode = str(sheet.cell(column=1, row=i).value)
    connectionName = str(sheet.cell(column=2, row=i).value)
    print(str(apCode) + " connect to " + str(connectionName))
    airportFoundInList = False
    for j in range(len(airports)): ##go through airports
        if apCode == airports[j].getCode(): #if new airport exists in working database
            print("    Found origin airport in database")
            airportFoundInList = True
            if not airports[j].hasConnection(connectionName): #if the connection doesn't already exist
                print("    Connection doesn't already exist")
                connectionExistsAsAirport = False
                for k in range(len(airports)):
                    if connectionName == airports[k].getCode():
                        print("    Found dest in database")
                        tempCon1 = Connection(airports[k])
                        tempCon2 = Connection(airports[j])
                        airports[j].addConnection(tempCon1)
                        airports[k].addConnection(tempCon2)
                        connectionExistsAsAirport = True
                        break
                if not connectionExistsAsAirport:
                    print("    Creating new airport and adding it")
                    tempAir = Airport(connectionName)
                    tempCon1 = Connection(tempAir)
                    tempCon2 = Connection(airports[j])
                    tempAir.addConnection(tempCon2)
                    airports[j].addConnection(tempCon1)
                    airports.append(tempAir)
            else:
                print("    Connection already exists")
            break
    if not airportFoundInList:
        print("    Origin airport not found in database, adding to list and creating connection")
        tempAir = Airport(apCode)
        foundDestinationInList = False
        for j in range(len(airports)):
            if connectionName == airports[j].getCode():
                print("    Found destination in database")
                foundDestinationInList = True
                tempCon1 = Connection(airports[j])
                tempCon2 = Connection(tempAir)
                tempAir.addConnection(tempCon1)
                airports[j].addConnection(tempCon2)
                break
        if not foundDestinationInList:
            print("    Dest not found in databse, creating and adding it")
            tempAirDest = Airport(connectionName)
            tempCon1 = Connection(tempAirDest)
            tempCon2 = Connection(tempAir)
            tempAirDest.addConnection(tempCon2)
            tempAir.addConnection(tempCon1)
            airports.append(tempAirDest)

        airports.append(tempAir)
    
bubbleSort(airports)

#search for duplicates
for i in range(len(airports)):
    for j in range(len(airports)):
        if airports[i].getCode() == airports[j].getCode() and i != j:
            print("Duplicate")

#add data from airportdata.xls
sheet2 = airInfoWb.get_sheet_by_name('airports')
for i in range(len(airports)):
    for j in range(1, sheet2.max_row + 1):
        airportCode = str(sheet2.cell(column=5, row=j).value)
        if airportCode == airports[i].getCode():
            airports[i].setName(str(sheet2.cell(column=2, row=j).value))
            airports[i].setCity(str(sheet2.cell(column=3, row=j).value))
            airports[i].setCountry(str(sheet2.cell(column=4, row=j).value))
            airports[i].setLatitude(float(sheet2.cell(column=7, row=j).value))
            airports[i].setLongitude(float(sheet2.cell(column=8, row=j).value))

r = 6371e3
#calculate distances
for i in range(len(airports)):
    tempConnections = airports[i].getConnections()
    sourceLat = airports[i].getLatitude()
    sourceLon = airports[i].getLongitude()
    for j in range(len(tempConnections)):
        psi1 = math.radians(sourceLat)
        destLat = tempConnections[j].getAirport().getLatitude()
        psi2 = math.radians(destLat)
        destLon = tempConnections[j].getAirport().getLongitude()
        print(str(destLat) + " " + str(destLon) + " " + str(sourceLat) + " " + str(sourceLon))
        deltaPsi = math.radians(destLat - sourceLat)
        deltaLambda = math.radians(destLon - sourceLon)
        a = math.sin(deltaPsi/2) * math.sin(deltaPsi/2) + math.cos(psi1) * math.cos(psi2) * math.sin(deltaLambda/2) * math.sin(deltaLambda/2)
        print(a)
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
        d = r * c
        tempConnections[j].setDistance(d)

#print all airports
for i in range(len(airports)):
    print("Airport: " + airports[i].getCode())
    print("    " + airports[i].getName())
    print("    " + airports[i].getCity())
    print("    " + airports[i].getCountry())
    print("    " + str(airports[i].getLatitude()))
    print("    " + str(airports[i].getLongitude()))
    tempConnections = airports[i].getConnections()
    for j in range(len(tempConnections)):
       print("        Connection to: " + tempConnections[j].getAirport().getCode() + " Distance: " + str(tempConnections[j].getDistance()))



rapp = RouteAppInfo(airports)
with open('../files/database.p', 'wb') as output:
    pickle.dump(rapp, output, pickle.HIGHEST_PROTOCOL)